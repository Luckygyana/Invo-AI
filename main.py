import os
import re
import sys
os.environ['CUDA_VISIBLE_DEVICES'] = '-1'
os.environ['OMP_THREAD_LIMIT'] = '1'
import cv2
import time
import argparse
import string
import xlsxwriter
# import keras_ocr 
import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment
from cfg import config as CONF
from src.extractStructure import ExtractStructure
import dateutil.parser as dparser
import datefinder
import pytesseract
import logging
from flask import Flask
from flask_cors import CORS
from database.db import initialize_db
from database.models import Job, initialize_default_config, Configuration
# alph = '$%@.,&():ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789/\\\n '
# # alph = string.digits + string.ascii_lowercase + string.ascii_uppercase
# recognizer = keras_ocr.recognition.Recognizer(alphabet=alph)
# print(recognizer.model.summary())
# pipeline = keras_ocr.pipeline.Pipeline()
# # sys.exit()
parser = argparse.ArgumentParser()
   
parser.add_argument('-i', '--input', required=True, help="input pdf")
parser.add_argument('-o', '--output', required=True, help="output xlsx")
parser.add_argument('--tess_exec', default=CONF.TESS_EXEC, help="tesseract_executable")
args = parser.parse_args()
if os.name == 'nt':
    pytesseract.pytesseract.tesseract_cmd = args.tess_exec



print(args.input)

inp = args.input
out = args.output
def convert(inp, out):
    t = time.time()
    structure = ExtractStructure(inp, recognizer=None, jobId=0)
    print('1')
    print(time.time() - t)
    t = time.time()
    structure.extract_images()
    print('2')
    print(time.time() - t)
    t = time.time()
    structure.bboxes_mapper()
    print('3')
    print(time.time() - t)
    t = time.time()
    print(structure.images)
    structure.row_keyword_mapper()

    print('4')
    print(time.time() - t)
    t = time.time()
    # img = np.full((2500, 2500, 3), 255)
    structure.detect_header()
    print('5')
    print(time.time() - t)
    t = time.time()
    # for key in structure.mapping_l2r.keys():
    #     print(structure.mapping_l2r[key].similarity_index)
    # print(structure.mapping_l2r)
    # print("rows:\n")
    alph = string.ascii_uppercase
    count = 0
    for key in structure.mapping_l2r.keys():
        li = structure.mapping_l2r[key].row_string
        # for s in li:
        #     print(s)
        structure.mapping_l2r[key].header_idx = np.argmax(structure.mapping_l2r[key].similarity_index)
        # print(idx)
        # print(structure.mapping_l2r[key].row_string[idx])
        structure.mapping_l2r[key].sheet = structure.mapping_l2r[key].row_string[structure.mapping_l2r[key].header_idx:]
        structure.mapping_l2r[key].info = structure.mapping_l2r[key].row_string[:structure.mapping_l2r[key].header_idx]
        structure.mapping_l2r[key].search_attributes()
        sheet = structure.mapping_l2r[key].sheet
        info = structure.mapping_l2r[key].info
        # print(structure.mapping_l2r[key].details)
        # print(structure.mapping_l2r[key].info)
        # print('\n\n\n\n')
        # print(structure.mapping_l2r[key].sheet)

        ##########################################################
        ########################################################
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 100
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50

        
        info = [['Seller State','XXXXXXXXXX'], ['Seller ID','XXXXXXXXXX'],['Seller Name','XXXXXXXXXX'],
            ['Seller Address','XXXXXXXXXXXXXXX'],['Seller GSTIN Number','XXXXXXXXX'],['Country of Origin', 'India'], 
            ['Currency', 'INR'],['Description', '']]


        _details = {
            'Invoice Number' : None,		
            'Invoice Date': None,	
            'Due Date' : None,
            'Total Invoice amount entered by WH operator' : None,		
            'Total Invoice Quantity entered by WH operator' : None,	
            'Total TCS Collected' : None,
            'Round Off Charges' : None,
            'PO Number' : None,
            'Invoice Items Total Amount' : None,	
            'Invoice Items total quantity' : None,
            'Buyer GSTIN Number' : None,
            'Ship to Address' : None,

        }

        details = structure.mapping_l2r[key].details
        print(details)
        for d_key in details.keys():
            if d_key == 'no':
                if ':' in details[d_key]:
                    _details['Invoice Number'] = details[d_key].split(':')[-1]
                else:
                    _details['Invoice Number'] = details[d_key]

            if d_key == 'date':
                if ':' in details[d_key]:
                    _details['Invoice Date'] = details[d_key].split(':')[-1]
                else:
                    _details['Invoice Date'] = details[d_key]

            if d_key == 'ship':
                if ':' in details[d_key]:
                    _details['Ship to Address'] = details[d_key].split(':')[-1]
                else:
                    _details['Ship to Address'] = details[d_key]

            if d_key == 'Due':
                # if ':' in details[d_key]:
                #     _details['Due Date'] = details[d_key].split(':')[-1]
                # else:
                #     _details['Due Date'] = details[d_key]
                # for _ in {'due', 'date'}:
                #     print(_details['Due Date'])
                #     if _ in _details['Due Date']:
                #         _details['Due Date'] = _details['Due Date'].replace(_, '')
                # dates = re.finditer("([0-9]{2}\.[0-9]{2}\.[0-9]{4})|([0-9]{2}\-[0-9]{2}\-[0-9]{4})|([0-9]{2}\/[0-9]{2}\/[0-9]{4})", details[d_key])
                # dates = [date.group() for date in dates if date]
                dates = datefinder.find_dates(details[d_key])
                dates = [date for date in dates if date]
                print(dates)
                try:
                    _details['Due Date'] = dates[0].strftime("%Y-%m-%d")
                except Exception as e:
                    print(e)
                    # sys.exit(0)


            if d_key == 'gst':
                # if ':' in details[d_key]:
                #     _details['Buyer GSTIN Number'] = details[d_key].split(':')[-1]
                # else:
                #     _details['Buyer GSTIN Number'] = details[d_key]
                gstin = re.search("([a-zA-Z0-9]{15})", details[d_key])
                try:
                    _details['Buyer GSTIN Number'] = gstin.group()
                except:
                    pass

            if d_key == 'po':
                if ':' in details[d_key]:
                    _details['PO Number'] = details[d_key].split(':')[-1]
                else:
                    _details['PO Number'] = details[d_key]



        print(_details)

        red_A = openpyxl.styles.colors.Color(rgb='00FF0000')
        fill_A = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=red_A)
        green_B = openpyxl.styles.colors.Color(rgb='00008000')
        fill_B = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green_B)
        blue_15 = openpyxl.styles.colors.Color(rgb='00596FF3')
        fill_15 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=blue_15)

        ws['A1'] = 'GST Invoice'
        ws['A1'].fill = fill_15
        # for i in range(1, 12):
        i = 1
        for dkey in _details.keys():
            i += 1
            ws['C' + str(i)] = dkey
            ws['D' + str(i)]= _details[dkey]

            ws['C' + str(i)].fill = fill_A
            ws['D' + str(i)].fill = fill_B

        for i in range(2, 10):
            for j in range(2):
                try:
                    col = alph[j]
                    idx = col + str(i)
                    # print(idx, i, j)
                    ws[idx] = info[i - 2][j]
                    if(col == 'A'):
                        ws[idx].fill = fill_A
                    if(col == 'B'):
                        ws[idx].fill = fill_B
                    ws[idx].alignment = Alignment(wrapText=True)
                except:
                    continue


        

        
        for i in range(1, len(sheet) + 1):
            for j in range(len(sheet[i - 1])):
                try:
                    col = alph[j]
                    idx = col + str(i + 15)
                    # print(idx, i, j)
                    ws[idx] = sheet[i - 1][j]
                    if(i == 1):
                        ws[idx].fill = fill_15
                    ws[idx].alignment = Alignment(wrapText=True)
                except:
                    continue
        output_location = os.path.join(args.output, re.split('\\\\|/', key)[-1].split('.')[0]+'.xlsx')
        print(args.output)
        print(output_location)
        wb.save(output_location)
        structure.result[structure.filename][count] = output_location

        count += 1
        #######################################################


        # for row in structure.mapping_l2r[key].rows:
        #     for box in row:
        #         x, y, w, h = box
        #         img = cv2.rectangle(img,(x,y),(x+w,y+h),(0, 255, 0),2)

        # print(structure.key_mapping_l2r[key])

    # plt.imshow(img)
    # plt.show()
    print('6')
    print(time.time() - t)
    t = time.time()

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)

    logger = logging.getLogger('Grid 2')

    UPLOAD_FOLDER = 'uploads'
    OUTPUT_FOLDER = 'output'
    ALLOWED_EXTENSIONS = set(['pdf'])

    app = Flask(__name__)
    CORS(app)

    app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
    app.config['MONGODB_DB'] = 'grid'
    # app.config['MONGODB_SETTINGS'] = {
    #     'host': os.environ['DB_PORT_27017_TCP_ADDR'],
    #     'port': 27017
    # }
    initialize_db(app)
    initialize_default_config()
    convert(inp=inp, out=out)